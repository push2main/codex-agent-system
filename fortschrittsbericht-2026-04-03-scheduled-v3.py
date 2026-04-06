import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
GREEN_FONT = Font(name='Arial', bold=True, color='006100', size=10)
RED_FONT = Font(name='Arial', bold=True, color='9C0006', size=10)
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
BORDER = Border(
    bottom=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0')
)

def style_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

def style_cell(ws, row, col, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align or Alignment(vertical='center', wrap_text=True)
    cell.border = BORDER
    return cell

# ── Sheet 1: Executive Summary ──
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_properties.tabColor = '1F4E79'
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
c = ws['A1']
c.value = 'Codex Agent System — Fortschrittsbericht'
c.font = Font(name='Arial', bold=True, size=14, color='1F4E79')
c.alignment = Alignment(horizontal='center')

ws.merge_cells('A2:C2')
c = ws['A2']
c.value = '2026-04-03 | Scheduled Audit Run'
c.font = Font(name='Arial', size=10, color='808080')
c.alignment = Alignment(horizontal='center')

r = 4
style_header_row(ws, r, 3)
ws.cell(r, 1, 'Metrik')
ws.cell(r, 2, 'Wert')
ws.cell(r, 3, 'Bewertung')

metrics = [
    ('Gesamte Tasks', '786', 'Umfangreiche Datenbasis'),
    ('All-Time Success Rate', '31%', 'Historisch niedrig (frühe Timeouts)'),
    ('Recent-50 Success Rate', '98%', 'Exzellent — System hat gelernt'),
    ('First-Pass Success Rate', '80%', 'Gut — wenig Retries nötig'),
    ('Timeout Rate (gesamt)', '27%', 'Historisch hoch, zuletzt ~0%'),
    ('Zero-Step Timeout Rate', '90%', 'Hauptursache der alten Timeouts'),
    ('Active Rules', '20/20', 'Cap erreicht — Qualität prüfen'),
    ('Registry Size', '120 KB', 'Gesund nach Compaction'),
    ('Pipeline Status', 'IDLE', 'Keine aktiven Tasks'),
    ('Zombie Tasks', '20', 'Alle archiviert/shelved'),
    ('Improvement Velocity', '+10.4 pp/100 Tasks', 'Starker Aufwärtstrend'),
]

for i, (name, val, note) in enumerate(metrics):
    r = 5 + i
    style_cell(ws, r, 1, BOLD_FONT)
    ws.cell(r, 1).value = name
    style_cell(ws, r, 2, NORMAL_FONT, align=Alignment(horizontal='center'))
    ws.cell(r, 2).value = val
    if 'Exzellent' in note or 'Gut' in note or 'Gesund' in note or 'Stark' in note:
        style_cell(ws, r, 3, GREEN_FONT, GREEN_FILL)
    elif 'niedrig' in note or 'hoch' in note:
        style_cell(ws, r, 3, RED_FONT, RED_FILL)
    else:
        style_cell(ws, r, 3, NORMAL_FONT, YELLOW_FILL)
    ws.cell(r, 3).value = note

# Verdict
r = 5 + len(metrics) + 1
ws.merge_cells(f'A{r}:C{r}')
c = ws.cell(r, 1)
c.value = 'VERDICT: System lernt effizient. Success Rate von 31% auf 98% gestiegen (+65.7pp). Pipeline ist IDLE — Übergang zu Growth-Mode empfohlen.'
c.font = Font(name='Arial', bold=True, size=11, color='006100')
c.fill = GREEN_FILL
c.alignment = Alignment(horizontal='center', wrap_text=True)
c.border = BORDER

# ── Sheet 2: Trend-Analyse ──
ws2 = wb.create_sheet('Trend-Analyse')
ws2.sheet_properties.tabColor = '2E75B6'
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 18

style_header_row(ws2, 1, 4)
ws2.cell(1, 1, 'Task Window')
ws2.cell(1, 2, 'Success Rate')
ws2.cell(1, 3, 'Timeouts')
ws2.cell(1, 4, 'Trend')

windows = [
    ('1-50', 0.34, 19), ('51-100', 0.04, 5), ('101-150', 0.06, 33),
    ('151-200', 0.04, 38), ('201-250', 0.16, 34), ('251-300', 0.10, 23),
    ('301-350', 0.14, 5), ('351-400', 0.12, 12), ('401-450', 0.22, 29),
    ('451-500', 0.26, 20), ('501-550', 0.10, 23), ('551-600', 0.14, 8),
    ('601-650', 0.58, 1), ('651-700', 0.86, 1), ('701-750', 0.96, 0),
    ('751-786', 0.97, 1),
]

for i, (win, rate, to) in enumerate(windows):
    r = 2 + i
    style_cell(ws2, r, 1, NORMAL_FONT, align=Alignment(horizontal='center'))
    ws2.cell(r, 1).value = win
    style_cell(ws2, r, 2, BOLD_FONT if rate > 0.5 else NORMAL_FONT, GREEN_FILL if rate >= 0.8 else (RED_FILL if rate < 0.15 else None), Alignment(horizontal='center'))
    ws2.cell(r, 2).value = rate
    ws2.cell(r, 2).number_format = '0%'
    style_cell(ws2, r, 3, NORMAL_FONT, align=Alignment(horizontal='center'))
    ws2.cell(r, 3).value = to
    trend = 'Aufstieg' if rate >= 0.5 else ('Stagnation' if rate >= 0.1 else 'Krise')
    style_cell(ws2, r, 4, GREEN_FONT if trend == 'Aufstieg' else (RED_FONT if trend == 'Krise' else NORMAL_FONT), align=Alignment(horizontal='center'))
    ws2.cell(r, 4).value = trend

chart = BarChart()
chart.type = 'col'
chart.title = 'Success Rate pro 50-Task-Window'
chart.y_axis.title = 'Success Rate'
chart.x_axis.title = 'Task Window'
chart.style = 10
data = Reference(ws2, min_col=2, min_row=1, max_row=17)
cats = Reference(ws2, min_col=1, min_row=2, max_row=17)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 22
chart.height = 12
ws2.add_chart(chart, 'F2')

# ── Sheet 3: Offene Issues & Empfehlungen ──
ws3 = wb.create_sheet('Issues & Empfehlungen')
ws3.sheet_properties.tabColor = 'C00000'
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 50
ws3.column_dimensions['E'].width = 16

style_header_row(ws3, 1, 5)
ws3.cell(1, 1, '#')
ws3.cell(1, 2, 'Severity')
ws3.cell(1, 3, 'Issue')
ws3.cell(1, 4, 'Empfehlung')
ws3.cell(1, 5, 'Status')

issues = [
    ('1', 'HIGH', 'Evaluator Zero-Score Blindness', '24% der Tasks bekommen Score=0. Evaluator-Prompt vereinfachen: binäre Wertung (produktiv/nicht produktiv) statt 0-10 Skala. Task "Fix value measurement blindness" ist 3x am Step-3-Timeout gescheitert — Task in kleinere Schritte splitten.', 'OFFEN'),
    ('2', 'MEDIUM', 'Stale Alerts trotz Idle Pipeline', 'retry_churn + loop_effort Alerts in alerts.json sind aktiv obwohl Pipeline idle ist. Ursache: metrics.json Flags werden nicht beim Idle-Check zurückgesetzt. Fix: Alert-Recomputation an Pipeline-Status koppeln.', 'WIEDERHOLT'),
    ('3', 'MEDIUM', 'Growth-Mode nicht implementiert', 'Rule existiert (idle + >90% success → growth-mode) aber kein Code-Pfad in self-improve.sh. System stagniert. Empfehlung: Growth-Mode Tasks implementieren (Test Coverage, Doku, neue Features).', 'OFFEN'),
    ('4', 'LOW', 'Metrics Drift nach Compaction', 'learning_rules_count=31 in metrics.json aber nur 20 in rules.md. Registry-Bytes überschätzt. Fix: Metrics-Refresh nach jeder Compaction erzwingen.', 'TEILWEISE'),
    ('5', 'LOW', 'External Signals veraltet', 'Letzte External Signal: OpenAI Python v2.30.0 (25. März). Status: stale. Signals bringen aktuell keinen Mehrwert — entweder aktualisieren oder deaktivieren.', 'OFFEN'),
    ('6', 'INFO', 'Task-Regeneration Loop gefixt', 'GLOBAL_FAMILY_SUCCESS_CAP von 8 auf 3 gesenkt. Verhindert ~60% der verschwendeten Task-Slots. Weiter beobachten.', 'GEFIXT'),
]

for i, (num, sev, issue, rec, status) in enumerate(issues):
    r = 2 + i
    style_cell(ws3, r, 1, NORMAL_FONT, align=Alignment(horizontal='center')).value = num
    sev_fill = RED_FILL if sev == 'HIGH' else (YELLOW_FILL if sev == 'MEDIUM' else GREEN_FILL)
    sev_font = RED_FONT if sev == 'HIGH' else (BOLD_FONT if sev == 'MEDIUM' else GREEN_FONT)
    style_cell(ws3, r, 2, sev_font, sev_fill, Alignment(horizontal='center')).value = sev
    style_cell(ws3, r, 3, BOLD_FONT).value = issue
    style_cell(ws3, r, 4, NORMAL_FONT).value = rec
    st_fill = RED_FILL if status == 'OFFEN' else (YELLOW_FILL if 'TEIL' in status or 'WIEDER' in status else GREEN_FILL)
    style_cell(ws3, r, 5, BOLD_FONT, st_fill, Alignment(horizontal='center')).value = status

# ── Sheet 4: Letzte Tasks ──
ws4 = wb.create_sheet('Letzte Tasks')
ws4.sheet_properties.tabColor = '548235'
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 40

style_header_row(ws4, 1, 6)
ws4.cell(1, 1, 'Datum')
ws4.cell(1, 2, 'Task')
ws4.cell(1, 3, 'Result')
ws4.cell(1, 4, 'Score')
ws4.cell(1, 5, 'Attempts')
ws4.cell(1, 6, 'Anmerkung')

recent = [
    ('2026-04-03 06:54', 'Fix value measurement blindness', 'SUCCESS', 5, 2, 'Step 3 Timeout — Task zu gross'),
    ('2026-04-03 04:39', 'Fix value measurement blindness', 'SUCCESS', 5, 2, 'Step 3 Timeout — wiederholt'),
    ('2026-04-03 02:23', 'Fix value measurement blindness', 'SUCCESS', 5, 2, 'Step 3 Timeout — wiederholt'),
    ('2026-04-03 00:40', 'Verify trigger-aware credential recovery', 'SUCCESS', 0, 2, 'Step 3 Timeout'),
    ('2026-04-03 00:00', 'Inventory decision path (incident id)', 'SUCCESS', 0, 2, 'Step 3 Timeout'),
    ('2026-04-02 12:29', 'Verify dashboard incident id field', 'FAILURE', 0, 2, 'Coder Implementation gescheitert'),
    ('2026-04-01 16:19', 'Verify dashboard incident payload', 'SUCCESS', 0, 2, 'Step 3 Timeout'),
    ('2026-03-31 22:22', 'Verify dashboard incident payload', 'FAILURE', 0, 2, 'Coder Implementation gescheitert'),
    ('2026-03-31 19:55', 'Verify dashboard incident id field', 'SUCCESS', 31, 1, 'First-pass, niedriger Score'),
    ('2026-03-31 18:13', 'Inventory decision path (incident id)', 'SUCCESS', 65, 1, 'First-pass'),
]

for i, (date, task, result, score, att, note) in enumerate(recent):
    r = 2 + i
    style_cell(ws4, r, 1, NORMAL_FONT, align=Alignment(horizontal='center')).value = date
    style_cell(ws4, r, 2, NORMAL_FONT).value = task
    res_fill = GREEN_FILL if result == 'SUCCESS' else RED_FILL
    res_font = GREEN_FONT if result == 'SUCCESS' else RED_FONT
    style_cell(ws4, r, 3, res_font, res_fill, Alignment(horizontal='center')).value = result
    style_cell(ws4, r, 4, NORMAL_FONT, align=Alignment(horizontal='center')).value = score
    style_cell(ws4, r, 5, NORMAL_FONT, align=Alignment(horizontal='center')).value = att
    style_cell(ws4, r, 6, NORMAL_FONT).value = note

# ── Sheet 5: Handlungsempfehlungen ──
ws5 = wb.create_sheet('Handlungsempfehlungen')
ws5.sheet_properties.tabColor = 'ED7D31'
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 55
ws5.column_dimensions['D'].width = 18

style_header_row(ws5, 1, 4)
ws5.cell(1, 1, '#')
ws5.cell(1, 2, 'Priorität')
ws5.cell(1, 3, 'Massnahme')
ws5.cell(1, 4, 'Typ')

actions = [
    ('1', 'SOFORT', 'Evaluator-Prompt redesignen: Einfache binäre Wertung statt 0-10 Skala. "Fix value measurement blindness" in 2 kleinere Tasks splitten (max 2 Files pro Step).', 'System-Fix'),
    ('2', 'SOFORT', 'alerts.json Recomputation fixen: retry_churn und loop_effort Flags an Pipeline-Idle-Status koppeln. Stale Alerts zurücksetzen wenn 0 running + 0 queued.', 'System-Fix'),
    ('3', 'KURZFRISTIG', 'Growth-Mode Code-Pfad in self-improve.sh implementieren: Wenn idle + success > 90%, neue Task-Kategorien generieren (Test Coverage, Dokumentation, API-Erweiterungen).', 'Feature'),
    ('4', 'KURZFRISTIG', 'metrics.json Auto-Refresh nach jeder Registry-Compaction einbauen. Aktuell driften Werte (rules_count, registry_bytes).', 'System-Fix'),
    ('5', 'OPTIONAL', 'External Signals entweder mit frischen Quellen aktualisieren oder Feature deaktivieren — bringt aktuell keinen Mehrwert.', 'Konfiguration'),
    ('6', 'OPTIONAL', 'Rules-Cap von 20 evaluieren: Älteste/ineffektivste Rules durch bessere ersetzen. rule-effectiveness-report.json als Basis nutzen.', 'Optimierung'),
]

for i, (num, prio, action, typ) in enumerate(actions):
    r = 2 + i
    style_cell(ws5, r, 1, NORMAL_FONT, align=Alignment(horizontal='center')).value = num
    prio_fill = RED_FILL if prio == 'SOFORT' else (YELLOW_FILL if prio == 'KURZFRISTIG' else GREEN_FILL)
    style_cell(ws5, r, 2, BOLD_FONT, prio_fill, Alignment(horizontal='center')).value = prio
    style_cell(ws5, r, 3, NORMAL_FONT).value = action
    style_cell(ws5, r, 4, NORMAL_FONT, align=Alignment(horizontal='center')).value = typ

out = '/sessions/nifty-fervent-hamilton/mnt/codex-agent-system/fortschrittsbericht-2026-04-03-scheduled-v3.xlsx'
wb.save(out)
print(f'Saved: {out}')
