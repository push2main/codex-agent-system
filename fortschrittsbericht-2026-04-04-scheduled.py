import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Color palette ---
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)
GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')
GOOD_FONT = Font(name='Arial', size=10, color='006100')
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
WARN_FONT = Font(name='Arial', size=10, color='9C6500')
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')
BAD_FONT = Font(name='Arial', size=10, color='9C0006')
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)
CENTER = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(vertical='top', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_row(ws, row, cols, font=BODY_FONT, fill=None, align=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = THIN_BORDER
        if fill: cell.fill = fill
        if align: cell.alignment = align

# ============================================================
# Sheet 1: Executive Summary
# ============================================================
ws = wb.active
ws.title = 'Executive Summary'
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50

ws.merge_cells('A1:C1')
ws['A1'] = 'Codex Agent System — Fortschrittsbericht 2026-04-04'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:C2')
ws['A2'] = 'Automatisch generiert | Pipeline-Status: IDLE'
ws['A2'].font = Font(name='Arial', italic=True, size=10, color='808080')
ws['A2'].alignment = Alignment(horizontal='center')

r = 4
headers = ['Metrik', 'Wert', 'Bewertung']
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)

metrics = [
    ('Tasks gesamt', 786, 'Solide Datenbasis'),
    ('All-time Success Rate', '31%', 'Durch Early Failures gedrückt'),
    ('Recent-50 Success Rate', '98%', 'Peak Performance, stabil'),
    ('Verbesserungs-Trajektorie', '34% → 97%', 'Klare, nachhaltige Verbesserung'),
    ('First-Pass Success', '57% (4/7)', 'Akzeptabel für Registry-Tasks'),
    ('Timeout-Rate', '27%', 'Historisch; kürzlich ~0%'),
    ('Aktive Regeln', '26 (dedupliziert)', 'Gesund, Eviction-Mechanismus aktiv'),
    ('Lernrate', '3.31 Regeln / 100 Tasks', 'Stabile Rule Discovery'),
    ('Registry-Größe', '124 KB / 512 KB', 'Weit unter Schwellwert'),
    ('Pipeline-Status', 'IDLE seit 03.04. 06:54', 'Kein aktiver Task'),
    ('Alerts', '0 aktiv', 'Alle bereinigt'),
    ('Zero-Step Timeouts', '90% der Timeouts', 'Planner-Cap bei 60s eingeführt'),
]

for i, (m, v, b) in enumerate(metrics):
    r += 1
    ws.cell(row=r, column=1, value=m)
    ws.cell(row=r, column=2, value=v)
    ws.cell(row=r, column=3, value=b)
    fill = None
    font = BODY_FONT
    if 'Peak' in b or 'Gesund' in b or 'bereinigt' in b or 'Weit unter' in b:
        fill, font = GOOD_FILL, GOOD_FONT
    elif 'gedrückt' in b or 'Historisch' in b:
        fill, font = WARN_FILL, WARN_FONT
    style_row(ws, r, 3, font=font, fill=fill)

# ============================================================
# Sheet 2: Trend-Analyse
# ============================================================
ws2 = wb.create_sheet('Trend-Analyse')
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 40

r = 1
ws2.cell(row=r, column=1, value='Window')
ws2.cell(row=r, column=2, value='Success Rate')
ws2.cell(row=r, column=3, value='Timeouts')
ws2.cell(row=r, column=4, value='Phase')
style_header(ws2, r, 4)

windows = [
    ('1-50', 0.34, 19, 'Frühphase — Viele Fehlschläge, System lernt'),
    ('51-100', 0.04, 5, 'Tiefpunkt — Komplexe Tasks, wenig Regeln'),
    ('101-150', 0.06, 33, 'Timeout-Krise — Planner-Overload'),
    ('151-200', 0.04, 38, 'Timeout-Krise — Peak Timeouts'),
    ('201-250', 0.16, 34, 'Erholung beginnt — Erste Regeln greifen'),
    ('251-300', 0.10, 23, 'Timeouts sinken, Review-Probleme'),
    ('301-350', 0.14, 5, 'Timeout-Fix wirkt (60s Cap)'),
    ('351-400', 0.12, 12, 'Stabilisierung'),
    ('401-450', 0.22, 29, 'Weiterer Timeout-Rückfall'),
    ('451-500', 0.26, 20, 'Langsame Verbesserung'),
    ('501-550', 0.10, 23, 'Rückschlag durch neue Task-Typen'),
    ('551-600', 0.14, 8, 'Recovery'),
    ('601-650', 0.58, 1, 'Durchbruch — Regelwerk greift'),
    ('651-700', 0.86, 1, 'Rapid Improvement'),
    ('701-750', 0.96, 0, 'Near-Perfect Execution'),
    ('751-786', 0.97, 1, 'Stabil auf Höchstniveau'),
]

for w, rate, to, phase in windows:
    r += 1
    ws2.cell(row=r, column=1, value=w)
    ws2.cell(row=r, column=2, value=rate)
    ws2.cell(row=r, column=2).number_format = '0%'
    ws2.cell(row=r, column=3, value=to)
    ws2.cell(row=r, column=4, value=phase)
    fill = GOOD_FILL if rate >= 0.80 else (WARN_FILL if rate >= 0.20 else BAD_FILL)
    font = GOOD_FONT if rate >= 0.80 else (WARN_FONT if rate >= 0.20 else BAD_FONT)
    style_row(ws2, r, 4, font=font, fill=fill)

# ============================================================
# Sheet 3: Tasks & Feasibility
# ============================================================
ws3 = wb.create_sheet('Tasks & Feasibility')
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 45

r = 1
for i, h in enumerate(['Status', 'Task-Titel', 'Projekt', 'Score', 'Feasibility-Bewertung'], 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, 5)

tasks = [
    ('completed', 'Verify dashboard incident payload coverage in smoke flow', 'superheld', 100, 'Erfolgreich — First-Pass'),
    ('completed', 'Inventory decision path for verify dashboard incident id', 'superheld', 100, 'Erfolgreich — First-Pass'),
    ('completed', 'Verify dashboard incident id field in smoke flow', 'superheld', 100, 'Erfolgreich — First-Pass'),
    ('completed', 'Verify trigger-aware credential recovery routing', 'superheld', 100, 'Erfolgreich — First-Pass'),
    ('completed', 'Fix value measurement blindness (v1)', 'superheld', 5, 'Teilweise — Score niedrig, Timeout bei Step 3'),
    ('completed', 'Fix value measurement blindness (v2)', 'superheld', 5, 'Teilweise — Score niedrig, Timeout bei Step 3'),
    ('completed', 'Self-learning audit v15-v16', 'codex-agent-system', '-', 'Erfolgreich — 3 Code-Fixes + Memory-Hygiene'),
    ('shelved', 'Verify dashboard incident payload coverage (retry)', 'superheld', 0, 'Zombie — 5+ Fehlversuche, korrekt geshelved'),
    ('shelved', 'Verify dashboard incident id field (retry)', 'superheld', 0, 'Zombie — Timeout-Wiederholer, korrekt geshelved'),
    ('shelved', 'Inventory decision path (retry)', 'superheld', 0, 'Zombie — Wiederholungs-Pattern erkannt'),
    ('shelved', 'Verify trigger-aware routing (retry)', 'superheld', 0, 'Zombie — Step 3 Timeout reproduzierbar'),
    ('shelved', 'Fix value measurement blindness (retry)', 'superheld', 0, 'Zombie — Scope zu groß für Budget'),
    ('pending_approval', 'Growth-mode Candidate (unbekannt)', 'codex-agent-system', '-', 'Blockiert durch Cooldown — Freigabe nötig'),
]

for status, title, proj, score, feas in tasks:
    r += 1
    ws3.cell(row=r, column=1, value=status)
    ws3.cell(row=r, column=2, value=title)
    ws3.cell(row=r, column=3, value=proj)
    ws3.cell(row=r, column=4, value=score)
    ws3.cell(row=r, column=5, value=feas)
    if status == 'completed' and score == 100:
        fill, font = GOOD_FILL, GOOD_FONT
    elif status == 'shelved':
        fill, font = BAD_FILL, BAD_FONT
    elif status == 'pending_approval':
        fill, font = WARN_FILL, WARN_FONT
    else:
        fill, font = WARN_FILL, WARN_FONT
    style_row(ws3, r, 5, font=font, fill=fill)
    ws3.cell(row=r, column=5).alignment = WRAP

# ============================================================
# Sheet 4: Empfehlungen
# ============================================================
ws4 = wb.create_sheet('Empfehlungen')
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 55
ws4.column_dimensions['D'].width = 14

r = 1
for i, h in enumerate(['Prio', 'Bereich', 'Empfehlung', 'Aufwand'], 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, 4)

recs = [
    (1, 'Pipeline reaktivieren', 'Pipeline ist seit >24h idle. Growth-Mode Cooldown manuell aufheben oder pending_approval Task freigeben, damit neue Tasks generiert werden.', 'Niedrig'),
    (2, 'Zero-Score Problem', 'Tasks mit status=success aber score=0 (z.B. Inventory-Tasks) verzerren Metriken. Evaluator-Clamp auf score>=5 wurde eingebaut (v15), aber ältere Daten sind betroffen. Historische Korrektur optional.', 'Mittel'),
    (3, 'Growth-Mode dynamisieren', 'Aktuell nur 3 statische Kandidaten. Dynamische Task-Generierung aus rules-candidate.md implementieren, damit Pipeline im Idle nicht stagniert.', 'Mittel'),
    (4, 'Timeout-Budget erhöhen', 'Step 3 Timeouts bei superheld-Tasks wiederholt. Entweder Step-Budget auf 120s erhöhen oder Tasks in kleinere Schritte splitten.', 'Mittel'),
    (5, 'Metrics-Drift Monitoring', 'validate-metrics.sh läuft nicht automatisch im Idle. In memory-sync.sh oder als Scheduled Task einbinden (alle 6h).', 'Niedrig'),
    (6, 'Zombie-Guard prüfen', '20 Zombie-Tasks, 166 verschwendete Slots. Guard greift, aber Prävention (früheres Shelving nach 3 statt 5 Fehlversuchen) würde Ressourcen sparen.', 'Niedrig'),
    (7, 'System stabil halten', 'Keine grundlegenden System-Modifikationen nötig. Regelwerk funktioniert, Lernrate ist gesund, Registry-Druck niedrig. Focus auf Task-Qualität statt System-Changes.', 'Kein'),
]

for prio, bereich, emp, aufw in recs:
    r += 1
    ws4.cell(row=r, column=1, value=prio)
    ws4.cell(row=r, column=2, value=bereich)
    ws4.cell(row=r, column=3, value=emp)
    ws4.cell(row=r, column=4, value=aufw)
    style_row(ws4, r, 4)
    ws4.cell(row=r, column=3).alignment = WRAP
    ws4.cell(row=r, column=1).alignment = CENTER

# ============================================================
# Sheet 5: Fazit
# ============================================================
ws5 = wb.create_sheet('Fazit')
ws5.column_dimensions['A'].width = 80

ws5.merge_cells('A1:A1')
ws5['A1'] = 'Gesamtbewertung'
ws5['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

fazit_lines = [
    '',
    'Das System lernt effizient und die Success Rate ist stabil auf 98% (Recent-50).',
    'Die Verbesserungstrajektorie von 34% auf 97% über 786 Tasks zeigt nachhaltiges Lernen.',
    '',
    'Bisherige Tasks sind umsetzbar:',
    '  • 4 von 7 Registry-Tasks erfolgreich im First-Pass (57%)',
    '  • 8 Tasks korrekt geshelved (Zombie-Guard funktioniert)',
    '  • 1 Task wartet auf Freigabe (Growth-Mode Cooldown)',
    '',
    'Modifikationen am System sind NICHT dringend nötig.',
    'Die Hauptaktion ist die Reaktivierung der Pipeline (Cooldown aufheben),',
    'damit der Growth-Mode neue Tasks generiert und die Pipeline nicht stagniert.',
    '',
    'Empfohlene Priorität:',
    '  1. Pipeline-Cooldown aufheben → Neue Tasks generieren lassen',
    '  2. Growth-Mode dynamisieren → Statische Kandidaten durch regelbasierte ersetzen',
    '  3. Step-Budget für superheld-Tasks anpassen → Weniger Timeouts',
    '',
    'Keine grundlegenden System-Änderungen erforderlich.',
]

for i, line in enumerate(fazit_lines, 2):
    ws5.cell(row=i, column=1, value=line)
    ws5.cell(row=i, column=1).font = BODY_FONT
    ws5.cell(row=i, column=1).alignment = WRAP

outpath = '/sessions/pensive-busy-sagan/mnt/codex-agent-system/fortschrittsbericht-2026-04-04-scheduled.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')
