from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN = "70AD47"
LIGHT_GREEN = "E2EFDA"
RED = "FF0000"
ORANGE = "ED7D31"
LIGHT_ORANGE = "FCE4D6"
YELLOW_BG = "FFF2CC"
WHITE = "FFFFFF"
GRAY = "D9D9D9"

header_font = Font(name="Arial", bold=True, color=WHITE, size=12)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
subheader_font = Font(name="Arial", bold=True, color=WHITE, size=10)
subheader_fill = PatternFill("solid", fgColor=MED_BLUE)
normal_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
green_font = Font(name="Arial", bold=True, color="228B22", size=10)
red_font = Font(name="Arial", bold=True, color=RED, size=10)
orange_font = Font(name="Arial", bold=True, color="D2691E", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
orange_fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
yellow_fill = PatternFill("solid", fgColor=YELLOW_BG)
blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        cell.border = thin_border

# ===== SHEET 1: EXECUTIVE SUMMARY =====
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = DARK_BLUE

for c in range(1, 7):
    ws.column_dimensions[get_column_letter(c)].width = [0, 30, 18, 18, 18, 18, 30][c]

r = 1
ws.merge_cells("A1:F1")
ws.cell(r, 1, "Codex Agent System — Fortschrittsbericht").font = Font(name="Arial", bold=True, size=16, color=DARK_BLUE)
ws.cell(r, 1).alignment = Alignment(horizontal="center")
r = 2
ws.merge_cells("A2:F2")
ws.cell(r, 1, "Automatisierter Audit — 2026-04-03").font = Font(name="Arial", size=11, color="666666")
ws.cell(r, 1).alignment = Alignment(horizontal="center")

r = 4
for c, h in enumerate(["Metrik", "Wert", "Bewertung", "Trend", "Ziel", "Kommentar"], 1):
    ws.cell(r, c, h)
style_range(ws, r, 1, 6, header_font, header_fill, Alignment(horizontal="center"))

data = [
    ["Gesamt Success Rate", "30%", "Kritisch", "↑ Steigend", ">50%", "Historisch niedrig wg. Early-Phase-Fehlern"],
    ["Recent-50 Success Rate", "98%", "Exzellent", "↑ Stabil hoch", ">90%", "Ziel weit übertroffen"],
    ["Q4 (letzte 132) Rate", "98%", "Exzellent", "↑ Stabil", ">85%", "Nachhaltiger Fortschritt"],
    ["First-Pass Success Rate", "79%", "Gut", "↑ Verbessert", ">70%", "Guter Erstversuch-Erfolg"],
    ["Timeout Rate", "27%", "Warnung", "↓ Sinkend", "<15%", "Historisch, recent nahezu 0"],
    ["Retry Classification", "100%", "Exzellent", "↑ Von 24% auf 100%", "100%", "Vollständig klassifiziert"],
    ["Registry Pressure", "242 KB", "OK", "→ Stabil", "<512 KB", "Unter Grenzwert"],
    ["Learning Rules", "17 von 20", "Gut", "→ Stabil", "≤20", "Nahe Maximum, qualitativ gut"],
    ["Active Alerts", "2", "Warnung", "→ Legacy-Alerts", "0", "retry_churn + loop_effort (historisch)"],
]
for i, row_data in enumerate(data):
    r = 5 + i
    for c, val in enumerate(row_data, 1):
        ws.cell(r, c, val).font = normal_font
        ws.cell(r, c).border = thin_border
        ws.cell(r, c).alignment = Alignment(horizontal="center" if c > 1 else "left", wrap_text=True)
    # Color the Bewertung column
    bew = row_data[2]
    if bew == "Exzellent":
        ws.cell(r, 3).font = green_font
        ws.cell(r, 3).fill = green_fill
    elif bew == "Gut":
        ws.cell(r, 3).font = Font(name="Arial", bold=True, color="228B22", size=10)
        ws.cell(r, 3).fill = green_fill
    elif bew == "Warnung":
        ws.cell(r, 3).font = orange_font
        ws.cell(r, 3).fill = orange_fill
    elif bew == "Kritisch":
        ws.cell(r, 3).font = red_font
        ws.cell(r, 3).fill = PatternFill("solid", fgColor="FFE0E0")

r = 15
ws.merge_cells(f"A{r}:F{r}")
ws.cell(r, 1, "FAZIT: Das System hat sich dramatisch verbessert. Die letzten 134 Tasks haben 98% Erfolgsquote.").font = Font(name="Arial", bold=True, size=11, color="228B22")
ws.cell(r, 1).fill = green_fill
ws.cell(r, 1).alignment = Alignment(horizontal="center")
for c in range(1,7): ws.cell(r,c).border = thin_border

# ===== SHEET 2: TREND ANALYSE =====
ws2 = wb.create_sheet("Trend Analyse")
ws2.sheet_properties.tabColor = MED_BLUE
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 25
ws2.column_dimensions['E'].width = 35

r = 1
for c, h in enumerate(["Window", "Success Rate", "Timeouts", "Phase", "Bewertung"], 1):
    ws2.cell(r, c, h)
style_range(ws2, r, 1, 5, header_font, header_fill, Alignment(horizontal="center"))

trends = [
    ("1-50", 0.34, 19, "Early Exploration", "Initiale Lernphase"),
    ("51-100", 0.04, 5, "Early Exploration", "Tiefpunkt — System lernt"),
    ("101-150", 0.06, 33, "Timeout-Krise", "Massive Timeout-Probleme"),
    ("151-200", 0.04, 38, "Timeout-Krise", "Schlimmste Phase"),
    ("201-250", 0.16, 34, "Erste Stabilisierung", "Langsame Erholung"),
    ("251-300", 0.10, 23, "Erste Stabilisierung", "Timeout-Reduktion beginnt"),
    ("301-350", 0.14, 5, "Timeout gelöst", "Timeouts drastisch reduziert"),
    ("351-400", 0.12, 12, "Plateau", "Stabile aber niedrige Rate"),
    ("401-450", 0.22, 29, "Verbesserung", "Timeout-Rückfall, aber höhere Rate"),
    ("451-500", 0.26, 20, "Verbesserung", "Aufwärtstrend setzt ein"),
    ("501-550", 0.10, 23, "Rückschlag", "Kurzfristiger Einbruch"),
    ("551-600", 0.14, 8, "Erholung", "Timeout unter Kontrolle"),
    ("601-650", 0.58, 1, "Durchbruch", "Massiver Qualitätssprung"),
    ("651-700", 0.86, 1, "High Performance", "System performt zuverlässig"),
    ("701-750", 0.96, 0, "Exzellenz", "Nahezu perfekt"),
    ("751-784", 0.97, 1, "Exzellenz", "Aktueller Stand — Spitzenleistung"),
]
for i, (w, rate, to, phase, bew) in enumerate(trends):
    r = 2 + i
    ws2.cell(r, 1, w).font = normal_font
    ws2.cell(r, 2, f"{rate:.0%}").font = normal_font
    ws2.cell(r, 3, to).font = normal_font
    ws2.cell(r, 4, phase).font = bold_font
    ws2.cell(r, 5, bew).font = normal_font
    for c in range(1, 6):
        ws2.cell(r, c).border = thin_border
        ws2.cell(r, c).alignment = Alignment(horizontal="center" if c < 4 else "left")
    if rate >= 0.90:
        ws2.cell(r, 2).fill = green_fill
        ws2.cell(r, 2).font = green_font
    elif rate >= 0.50:
        ws2.cell(r, 2).fill = yellow_fill
    elif rate < 0.10:
        ws2.cell(r, 2).fill = PatternFill("solid", fgColor="FFE0E0")
        ws2.cell(r, 2).font = red_font

# ===== SHEET 3: TASK STATUS =====
ws3 = wb.create_sheet("Tasks Aktuell")
ws3.sheet_properties.tabColor = GREEN
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 40

r = 1
for c, h in enumerate(["Task ID / Titel", "Status", "Kategorie", "Impact", "Effort", "Empfehlung"], 1):
    ws3.cell(r, c, h)
style_range(ws3, r, 1, 6, header_font, header_fill, Alignment(horizontal="center"))

tasks = [
    ("task-002: Unit test clamp_prompt_context 4k", "shelved", "testing", 7, 1, "REAKTIVIEREN — niedrig Effort, hoher Wert"),
    ("task-003: Unit test classify_retry_failure", "shelved", "testing", 7, 1, "REAKTIVIEREN — wichtig für Retry-Qualität"),
    ("task-004: Learner dedup threshold Kommentar", "shelved", "code_quality", 6, 1, "REAKTIVIEREN — triviale Doku-Aufgabe"),
    ("task-001: Review OpenAI Python v2.30.0", "shelved", "code_quality", 6, 2, "VERWERFEN — nicht relevant für System"),
    ("task-005: System-work buffer bei Queue-Drain", "shelved", "stability", 8, 2, "PRÜFEN — 324x Zombie, Grundansatz ändern"),
    ("task-006: Planner MAX_STEP_CHARS Kommentar", "completed", "code_quality", 5, 1, "ERLEDIGT"),
    ("task-007: Test planner step < 600 chars", "completed", "testing", 7, 1, "ERLEDIGT"),
    ("task-008: Learner 65% threshold Kommentar", "completed", "code_quality", 4, 1, "ERLEDIGT"),
    ("task-009: Inventory cap pre-step planning", "shelved", "learning", 0, 0, "VERWERFEN — Zombie-Kandidat"),
    ("task-010: Reduce timeout rate", "shelved", "performance", 6, 3, "VERWERFEN — Problem ist bereits gelöst (0 Timeouts recent)"),
    ("task-011: Improve retry success rate", "completed", "stability", 0, 0, "ERLEDIGT"),
]
for i, (title, status, cat, impact, effort, rec) in enumerate(tasks):
    r = 2 + i
    ws3.cell(r, 1, title).font = normal_font
    ws3.cell(r, 2, status).font = bold_font
    ws3.cell(r, 3, cat).font = normal_font
    ws3.cell(r, 4, impact).font = normal_font
    ws3.cell(r, 5, effort).font = normal_font
    ws3.cell(r, 6, rec).font = normal_font
    for c in range(1, 7):
        ws3.cell(r, c).border = thin_border
        ws3.cell(r, c).alignment = Alignment(wrap_text=True)
    if status == "completed":
        ws3.cell(r, 2).fill = green_fill
        ws3.cell(r, 2).font = green_font
    elif status == "shelved":
        ws3.cell(r, 2).fill = orange_fill
    if "REAKTIVIEREN" in rec:
        ws3.cell(r, 6).fill = yellow_fill
        ws3.cell(r, 6).font = bold_font
    elif "VERWERFEN" in rec:
        ws3.cell(r, 6).fill = PatternFill("solid", fgColor="FFE0E0")

# ===== SHEET 4: ZOMBIES =====
ws4 = wb.create_sheet("Zombie Tasks")
ws4.sheet_properties.tabColor = "FF0000"
ws4.column_dimensions['A'].width = 70
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 40

r = 1
for c, h in enumerate(["Zombie Task Titel", "Failures", "Empfehlung"], 1):
    ws4.cell(r, c, h)
style_range(ws4, r, 1, 3, header_font, PatternFill("solid", fgColor="C00000"), Alignment(horizontal="center"))

zombies = [
    ("Keep an executable system-work buffer when queue drains", 324, "PERMANENT SPERREN — unausführbar in aktueller Architektur"),
    ("Inventory decision path: improve first-pass success rate", 20, "ARCHIVIEREN — First-Pass ist jetzt bei 79%"),
    ("Inventory decision path: recover stale pipeline", 13, "ARCHIVIEREN — Pipeline ist nicht mehr stale"),
    ("Tighten mobile dashboard into enterprise control surface", 11, "SPERREN — zu vage, manuell planen"),
    ("Reduce timeout rate", 11, "ARCHIVIEREN — Timeout-Problem ist gelöst"),
    ("Detect retry churn and queue starvation", 6, "ARCHIVIEREN — Retry churn ist unter Kontrolle"),
    ("Improve first-pass success rate", 5, "ARCHIVIEREN — Ziel bereits erreicht (79%)"),
    ("Break retry churn", 5, "ARCHIVIEREN — Churn ist eingedämmt"),
    ("Cap pre-step planning budget", 5, "ARCHIVIEREN — Zero-step timeouts bei 0 recent"),
]
for i, (title, fails, rec) in enumerate(zombies):
    r = 2 + i
    ws4.cell(r, 1, title).font = normal_font
    ws4.cell(r, 2, fails).font = red_font
    ws4.cell(r, 3, rec).font = bold_font
    for c in range(1, 4):
        ws4.cell(r, c).border = thin_border
        ws4.cell(r, c).alignment = Alignment(wrap_text=True)
    ws4.cell(r, 1).fill = PatternFill("solid", fgColor="FFE0E0")

r += 2
ws4.merge_cells(f"A{r}:C{r}")
ws4.cell(r, 1, "DRINGEND: Alle 9 Zombie-Tasks sollten endgültig aus der Registry entfernt werden. Sie verursachen 400+ vergebliche Retry-Zyklen.").font = Font(name="Arial", bold=True, size=10, color=RED)

# ===== SHEET 5: EMPFEHLUNGEN =====
ws5 = wb.create_sheet("Empfehlungen")
ws5.sheet_properties.tabColor = "70AD47"
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 40
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 50

r = 1
for c, h in enumerate(["Prio", "Maßnahme", "Aufwand", "Impact", "Begründung"], 1):
    ws5.cell(r, c, h)
style_range(ws5, r, 1, 5, header_font, PatternFill("solid", fgColor="228B22"), Alignment(horizontal="center"))

recs = [
    ("1", "Zombie-Tasks permanent archivieren (9 Stk.)", "Minimal", "Hoch", "324+ vergebliche Versuche eliminieren, Registry-Druck senken"),
    ("2", "Shelved Tasks 002/003/004 reaktivieren", "Niedrig", "Mittel", "3 Quick-Win-Tasks mit Effort=1, stärken Testabdeckung"),
    ("3", "Alert retry_churn + loop_effort zurücksetzen", "Minimal", "Mittel", "Legacy-Alerts spiegeln nicht mehr den aktuellen Zustand"),
    ("4", "Obsolete Tasks 001/009/010 verwerfen", "Minimal", "Niedrig", "OpenAI-Review irrelevant, Timeout/Planning bereits gelöst"),
    ("5", "Self-Improve Cooldown prüfen", "Niedrig", "Mittel", "Cooldown blockiert aktuell jede Selbstverbesserung"),
    ("6", "External Signals aktualisieren", "Niedrig", "Niedrig", "Letzte Aktualisierung: 26.03. — 8 Tage stale"),
    ("7", "Archive kompaktieren", "Mittel", "Niedrig", "1103 Tasks im Archiv, davon 375 shelved — aufräumen"),
]
for i, (prio, action, effort, impact, reason) in enumerate(recs):
    r = 2 + i
    ws5.cell(r, 1, prio).font = bold_font
    ws5.cell(r, 2, action).font = bold_font
    ws5.cell(r, 3, effort).font = normal_font
    ws5.cell(r, 4, impact).font = normal_font
    ws5.cell(r, 5, reason).font = normal_font
    for c in range(1, 6):
        ws5.cell(r, c).border = thin_border
        ws5.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    if prio in ("1", "2"):
        ws5.cell(r, 1).fill = PatternFill("solid", fgColor="FFE0E0")
    elif prio in ("3", "4"):
        ws5.cell(r, 1).fill = yellow_fill

outpath = "/sessions/friendly-vigilant-gauss/mnt/codex-agent-system/fortschrittsbericht-2026-04-03-auto.xlsx"
wb.save(outpath)
print(f"Saved to {outpath}")
