import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# --- Colors ---
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
GREEN_FONT = Font(name='Arial', bold=True, size=10, color='006100')
RED_FONT = Font(name='Arial', bold=True, size=10, color='9C0006')
ORANGE_FONT = Font(name='Arial', bold=True, size=10, color='BF6900')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, font=None, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# ============ SHEET 1: Executive Summary ============
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_properties.tabColor = '1F4E79'

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 55

ws.merge_cells('A1:C1')
ws['A1'] = 'Codex Agent System — Fortschrittsbericht'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:C2')
ws['A2'] = f'Automatischer Scheduled Report — {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

r = 4
ws.cell(r, 1, 'Metrik').font = SECTION_FONT
ws.cell(r, 1).fill = SECTION_FILL
ws.cell(r, 2, 'Wert').font = SECTION_FONT
ws.cell(r, 2).fill = SECTION_FILL
ws.cell(r, 3, 'Bewertung').font = SECTION_FONT
ws.cell(r, 3).fill = SECTION_FILL
style_header(ws, r, 3)

data = [
    ('Pipeline-Status', 'IDLE', 'Keine aktiven, gequeueten oder approved Tasks'),
    ('Gesamt-Tasks', '786', 'Über beide Projekte (codex-agent-system + superheld)'),
    ('All-Time Success Rate', '31%', 'Niedrig wg. früher Phase — irreführend ohne Kontext'),
    ('Recent-50 Success Rate', '98%', 'Exzellent — System hat sich stabilisiert'),
    ('First-Pass Success Rate', '57% (global) / 80% (CLAUDE.md)', 'Gut, aber Verbesserungspotenzial bei first-pass'),
    ('Timeout Rate', '27% (all-time)', 'Historisch hoch, zuletzt 0-1 Timeouts pro 50er-Window'),
    ('Improvement Trend', '+65.7 pp', 'Massive Verbesserung von Fenster 1→16'),
    ('Active Alerts', '2 (retry_churn + loop_effort)', 'Stale Alerts — Pipeline ist idle, sollten gelöscht werden'),
    ('Registry Size', '142 KB / 15 Tasks', 'Unter Druckgrenze (512 KB), gesund'),
    ('Learned Rules', '20/20 (cap erreicht)', '31 in rules.md — Diskrepanz zu Cap, Bereinigung nötig'),
    ('Zombie Tasks', '20 (166 wasted slots)', 'Archiviert, kein aktiver Schaden mehr'),
    ('Self-Improve Status', 'cooldown_active', 'Blockiert neue Taskgenerierung — beabsichtigt'),
]

for i, (metric, value, assessment) in enumerate(data):
    r = 5 + i
    ws.cell(r, 1, metric).font = BOLD_FONT
    ws.cell(r, 2, value).font = NORMAL_FONT
    ws.cell(r, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(r, 3, assessment).font = NORMAL_FONT
    if 'Exzellent' in assessment or 'Massive' in assessment:
        ws.cell(r, 2).fill = GREEN_FILL
        ws.cell(r, 2).font = GREEN_FONT
    elif 'Niedrig' in assessment or 'hoch' in assessment or 'Stale' in assessment:
        ws.cell(r, 2).fill = YELLOW_FILL
        ws.cell(r, 2).font = ORANGE_FONT
    style_row(ws, r, 3)

# ============ SHEET 2: Trend Analysis ============
ws2 = wb.create_sheet('Trend-Analyse')
ws2.sheet_properties.tabColor = '2E75B6'
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 50

ws2.merge_cells('A1:D1')
ws2['A1'] = 'Success Rate nach 50er-Fenstern'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

r = 3
for c, h in enumerate(['Fenster', 'Success Rate', 'Timeouts', 'Trend-Kommentar'], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 4)

windows = [
    ('1-50', 0.34, 19, 'Frühphase, viel Trial-and-Error'),
    ('51-100', 0.04, 5, 'Tiefpunkt — System lernt noch'),
    ('101-150', 0.06, 33, 'Timeout-Explosion durch oversized context'),
    ('151-200', 0.04, 38, 'Timeout-Peak — MAX_PROMPT_CONTEXT fix ausstehend'),
    ('201-250', 0.16, 34, 'Erste Besserung nach Context-Clamp'),
    ('251-300', 0.10, 23, 'Timeout-Reduktion beginnt'),
    ('301-350', 0.14, 5, 'Deutliche Timeout-Senkung'),
    ('351-400', 0.12, 12, 'Plateau — Review-Rejection-Problem entdeckt'),
    ('401-450', 0.22, 29, 'Timeout-Rückfall bei komplexen Tasks'),
    ('451-500', 0.26, 20, 'Verbesserung durch File-Count-Splitting'),
    ('501-550', 0.10, 23, 'Rückschlag — Zombie-Tasks entdeckt'),
    ('551-600', 0.14, 8, 'Zombie-Guard + Cooldown eingeführt'),
    ('601-650', 0.58, 1, 'DURCHBRUCH — fokussierte Smoke-Test-Tasks'),
    ('651-700', 0.86, 1, 'Rapid Improvement — Dashboard-Payload-Tasks'),
    ('701-750', 0.96, 0, 'Near-perfekt — System stabil'),
    ('751-786', 0.97, 1, 'Plateau auf hohem Niveau'),
]

for i, (w, rate, to, comment) in enumerate(windows):
    r = 4 + i
    ws2.cell(r, 1, w).font = NORMAL_FONT
    ws2.cell(r, 1).alignment = Alignment(horizontal='center')
    ws2.cell(r, 2, f'{rate:.0%}').font = NORMAL_FONT
    ws2.cell(r, 2).alignment = Alignment(horizontal='center')
    if rate >= 0.80:
        ws2.cell(r, 2).fill = GREEN_FILL
        ws2.cell(r, 2).font = GREEN_FONT
    elif rate <= 0.10:
        ws2.cell(r, 2).fill = RED_FILL
        ws2.cell(r, 2).font = RED_FONT
    elif rate <= 0.30:
        ws2.cell(r, 2).fill = YELLOW_FILL
    ws2.cell(r, 3, to).font = NORMAL_FONT
    ws2.cell(r, 3).alignment = Alignment(horizontal='center')
    ws2.cell(r, 4, comment).font = NORMAL_FONT
    style_row(ws2, r, 4)

# ============ SHEET 3: Probleme & Empfehlungen ============
ws3 = wb.create_sheet('Probleme & Empfehlungen')
ws3.sheet_properties.tabColor = 'ED7D31'
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 55
ws3.column_dimensions['E'].width = 55

ws3.merge_cells('A1:E1')
ws3['A1'] = 'Identifizierte Probleme & Empfohlene Maßnahmen'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

r = 3
for c, h in enumerate(['#', 'Problem', 'Severity', 'Analyse', 'Empfehlung'], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 5)

problems = [
    (1, 'Regeneration Cycle (Hauptproblem)', 'HIGH',
     'Weakness-Detector erkennt wiederholt dieselben Milestones aus spec.md als fehlend, obwohl die Assertions bereits existieren. 4 Task-Familien liefen 95x mit identischem Ergebnis.',
     'Weakness-Detector muss prüfen ob Assertions im Codebase existieren BEVOR neue Tasks generiert werden. Milestone als "resolved" markieren wenn Assertion bereits passt.'),
    (2, 'Stale Alerts (retry_churn + loop_effort)', 'MEDIUM',
     'alerts.json zeigt 2 aktive Alerts obwohl Pipeline idle ist (0 running, 0 queued). Alerts wurden bei letztem Audit-Run nicht zurückgesetzt.',
     'Alerts in alerts.json zurücksetzen: active=false, alerts=[]. Regel existiert bereits (verify pipeline idle before clearing).'),
    (3, 'Rules Cap Diskrepanz', 'LOW',
     'CLAUDE.md sagt 20/20 Rules, aber rules.md enthält 21 Rules. Learner-Dedup funktioniert, aber Cap-Enforcement ist undicht.',
     'rules.md auf 20 trimmen. Learner prüfen ob merge+dedup korrekt vor Cap-Check läuft.'),
    (4, 'Score=0 bei erfolgreichen Tasks', 'MEDIUM',
     '3x "Fix value measurement blindness" Tasks liefen (score=5), aber das Kernproblem bleibt: Evaluator bewertet viele Tasks mit 0 obwohl sie erfolgreich sind.',
     'Evaluator-Prompt anpassen: Score basiert auf tatsächlichem Code-Delta, nicht nur auf Milestone-Completion. Minimum-Score für erfolgreiche Tasks auf 10 setzen.'),
    (5, 'cooldown_active blockiert Growth-Mode', 'LOW',
     'Pipeline idle + recent_success_rate > 0.95 → cooldown_active. CLAUDE.md-Regel sagt "shift to growth-mode" — aber cooldown blockiert.',
     'Growth-Mode-Regel höhere Priorität als cooldown geben. Wenn pipeline idle UND success > 0.90 → Growth-Tasks generieren (Doku, Test-Coverage, Capability-Expansion).'),
    (6, 'metrics.json Inkonsistenz', 'LOW',
     'metrics.json: task_registry_total=15, aber self-improve-run.json sagt "registry_count_mismatch". first_pass_success_rate=0.57 vs 1.0 je nach Quelle.',
     'Metriken-Refresh erzwingen nach jedem Compaction-Lauf. Einheitliche Quelle für first_pass_success definieren.'),
]

for i, (num, prob, sev, analysis, rec) in enumerate(problems):
    r = 4 + i
    ws3.cell(r, 1, num).font = NORMAL_FONT
    ws3.cell(r, 1).alignment = Alignment(horizontal='center')
    ws3.cell(r, 2, prob).font = BOLD_FONT
    ws3.cell(r, 3, sev).font = BOLD_FONT
    ws3.cell(r, 3).alignment = Alignment(horizontal='center')
    if sev == 'HIGH':
        ws3.cell(r, 3).fill = RED_FILL
        ws3.cell(r, 3).font = RED_FONT
    elif sev == 'MEDIUM':
        ws3.cell(r, 3).fill = YELLOW_FILL
        ws3.cell(r, 3).font = ORANGE_FONT
    else:
        ws3.cell(r, 3).fill = GREEN_FILL
    ws3.cell(r, 4, analysis).font = NORMAL_FONT
    ws3.cell(r, 5, rec).font = NORMAL_FONT
    style_row(ws3, r, 5)

# ============ SHEET 4: Provider Performance ============
ws4 = wb.create_sheet('Provider Performance')
ws4.sheet_properties.tabColor = '70AD47'
ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 16
ws4.column_dimensions['G'].width = 40

ws4.merge_cells('A1:G1')
ws4['A1'] = 'Provider Performance nach Kategorie (Claude vs Codex)'
ws4['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

r = 3
for c, h in enumerate(['Kategorie', 'Claude Rate', 'Claude Tasks', 'Codex Rate', 'Codex Tasks', 'Besser', 'Routing-Empfehlung'], 1):
    ws4.cell(r, c, h)
style_header(ws4, r, 7)

providers = [
    ('auth', 0.0, 3, 0.72, 29, 'Codex', 'Codex klar besser — Routing korrekt'),
    ('code_quality', 0.0, 5, 0.11, 19, 'Codex', 'Beide schwach — Tasks zu komplex?'),
    ('general', 0.19, 54, 0.25, 142, 'Codex', 'Codex leicht besser — Routing ok'),
    ('infra', 0.16, 32, 0.40, 92, 'Codex', 'Codex deutlich besser'),
    ('learning', 0.08, 12, 0.12, 34, 'Codex', 'Beide niedrig — Task-Qualität prüfen'),
    ('testing', 0.36, 14, 0.57, 7, 'Codex', 'Codex besser, aber wenig Daten'),
    ('ui', 0.19, 78, 0.41, 244, 'Codex', 'Codex klar besser — Claude für UI ineffektiv'),
]

for i, (cat, cr, ct, xr, xt, better, rec) in enumerate(providers):
    r = 4 + i
    ws4.cell(r, 1, cat).font = BOLD_FONT
    ws4.cell(r, 2, f'{cr:.0%}').font = NORMAL_FONT
    ws4.cell(r, 2).alignment = Alignment(horizontal='center')
    ws4.cell(r, 3, ct).font = NORMAL_FONT
    ws4.cell(r, 3).alignment = Alignment(horizontal='center')
    ws4.cell(r, 4, f'{xr:.0%}').font = NORMAL_FONT
    ws4.cell(r, 4).alignment = Alignment(horizontal='center')
    ws4.cell(r, 5, xt).font = NORMAL_FONT
    ws4.cell(r, 5).alignment = Alignment(horizontal='center')
    ws4.cell(r, 6, better).font = BOLD_FONT
    ws4.cell(r, 6).alignment = Alignment(horizontal='center')
    ws4.cell(r, 7, rec).font = NORMAL_FONT
    if xr > cr:
        ws4.cell(r, 4).fill = GREEN_FILL
    elif cr > xr:
        ws4.cell(r, 2).fill = GREEN_FILL
    style_row(ws4, r, 7)

# ============ SHEET 5: Gesamtbewertung ============
ws5 = wb.create_sheet('Gesamtbewertung')
ws5.sheet_properties.tabColor = 'FFC000'
ws5.column_dimensions['A'].width = 80

ws5.merge_cells('A1:A1')
ws5['A1'] = 'Gesamtbewertung & Fazit'
ws5['A1'].font = Font(name='Arial', bold=True, size=16, color='1F4E79')

texts = [
    ('GESAMTURTEIL: System funktioniert — Modifikationen notwendig, aber keine grundlegenden Umbauten', SECTION_FONT, SECTION_FILL),
    ('', NORMAL_FONT, None),
    ('FORTSCHRITT:', BOLD_FONT, None),
    ('Das Codex Agent System zeigt eine massive Verbesserung von 4% auf 97% Success Rate über die letzten 400 Tasks. Die Lernkurve ist real — das System hat aus Fehlern gelernt, Rules akkumuliert und die Pipeline stabilisiert. Die Recent-50 Rate von 98% zeigt, dass das aktuelle Regelwerk funktioniert.', NORMAL_FONT, None),
    ('', NORMAL_FONT, None),
    ('SIND DIE TASKS UMSETZBAR?', BOLD_FONT, None),
    ('Ja, die aktuellen Task-Patterns (fokussierte Smoke-Test-Assertions, Baseline-Verification-Guards) sind erfolgreich und reproduzierbar. Das Problem ist nicht die Umsetzbarkeit, sondern die Regeneration: identische Tasks werden wiederholt ausgeführt ohne Mehrwert.', NORMAL_FONT, None),
    ('', NORMAL_FONT, None),
    ('NOTWENDIGE MODIFIKATIONEN (priorisiert):', BOLD_FONT, None),
    ('1. [HIGH] Weakness-Detector fixen: Existenz-Check für Assertions vor Task-Generierung einbauen. Dies ist das Hauptproblem — 95 von 100 letzten Tasks waren unnötige Wiederholungen.', NORMAL_FONT, None),
    ('2. [MEDIUM] Stale Alerts bereinigen: retry_churn und loop_effort in alerts.json zurücksetzen (Pipeline ist idle).', NORMAL_FONT, None),
    ('3. [MEDIUM] Evaluator-Scoring fixen: Erfolgreiche Tasks sollten nie Score=0 bekommen. Minimum auf 10 setzen.', NORMAL_FONT, None),
    ('4. [LOW] Growth-Mode aktivieren: Wenn Pipeline idle + success > 90%, automatisch Doku/Test/Expansion-Tasks generieren statt idle zu bleiben.', NORMAL_FONT, None),
    ('5. [LOW] Rules-Cap enforcing: rules.md auf 20 trimmen, Dedup-Check vor Cap stärken.', NORMAL_FONT, None),
    ('', NORMAL_FONT, None),
    ('KEINE MODIFIKATION NÖTIG:', BOLD_FONT, None),
    ('Provider-Routing (Codex dominiert korrekt), Task-Registry-Size (gesund bei 142KB), Timeout-Handling (von 38 auf 0-1 pro Fenster), Retry-Classification (100% Coverage).', NORMAL_FONT, None),
]

for i, (text, font, fill) in enumerate(texts):
    r = 3 + i
    ws5.cell(r, 1, text).font = font
    if fill:
        ws5.cell(r, 1).fill = fill
    ws5.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')

out = '/sessions/gracious-modest-mayer/mnt/codex-agent-system/fortschrittsbericht-2026-04-03-scheduled-v2.xlsx'
wb.save(out)
print(f'Saved to {out}')
