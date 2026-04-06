import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Colors & Styles ---
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
SECTION_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')
OK_FILL = PatternFill('solid', fgColor='C6EFCE')
OK_FONT = Font(name='Arial', color='006100', size=10)
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
WARN_FONT = Font(name='Arial', color='9C6500', size=10)
BAD_FILL = PatternFill('solid', fgColor='FFC7CE')
BAD_FONT = Font(name='Arial', color='9C0006', size=10)
NORMAL = Font(name='Arial', size=10)
BOLD = Font(name='Arial', size=10, bold=True)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(bottom=THIN)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_section(ws, row, cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL

def status_cell(ws, row, col, value, status):
    cell = ws.cell(row=row, column=col, value=value)
    if status == 'ok':
        cell.fill, cell.font = OK_FILL, OK_FONT
    elif status == 'warn':
        cell.fill, cell.font = WARN_FILL, WARN_FONT
    elif status == 'bad':
        cell.fill, cell.font = BAD_FILL, BAD_FONT
    else:
        cell.font = NORMAL

# ============ SHEET 1: Summary ============
ws = wb.active
ws.title = 'Summary'
ws.sheet_properties.tabColor = '1F4E79'

for c, w in enumerate([40, 25, 20, 45], 1):
    ws.column_dimensions[get_column_letter(c)].width = w

r = 1
ws.merge_cells('A1:D1')
title = ws.cell(row=1, column=1, value='Codex Agent System — Fortschrittsbericht')
title.font = Font(name='Arial', bold=True, size=14, color='1F4E79')
title.alignment = Alignment(horizontal='center')

r = 2
ws.merge_cells('A2:D2')
ws.cell(row=2, column=1, value='Stand: 2026-04-03 | Pipeline: IDLE | 786 Tasks total').font = Font(name='Arial', size=10, italic=True, color='666666')
ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

r = 4
style_section(ws, r, 4, 'Kern-KPIs')
r = 5
style_header(ws, r, 4)
for c, h in enumerate(['Metrik', 'Wert', 'Bewertung', 'Kommentar'], 1):
    ws.cell(row=r, column=c, value=h)

kpis = [
    ('All-Time Success Rate', '31%', 'ok', 'Baseline — durch frühe Fehler gedrückt'),
    ('Recent-50 Success Rate', '98%', 'ok', 'Peak-Performance, stabil seit ~100 Tasks'),
    ('First-Pass Success (recent)', '84%', 'ok', 'Gut — 41/49 beim 1. Versuch erfolgreich'),
    ('Timeout-Rate (gesamt)', '27%', 'warn', 'Historisch hoch, aber aktuell nahe 0'),
    ('Timeout-Rate (recent-50)', '~2%', 'ok', '1 Timeout in letzten 36 Tasks'),
    ('Rules aktiv', '20/20', 'warn', 'Am Cap — kein Eviction-Mechanismus'),
    ('Registry Größe', '75 KB', 'ok', 'Gesund, weit unter 512 KB Schwelle'),
    ('Pipeline Status', 'IDLE', 'ok', '0 running, 0 queued, 0 approved'),
    ('Zombie Tasks', '20 archiviert', 'ok', 'Korrekt aussortiert'),
    ('Metrics Drift', 'Wiederkehrend', 'warn', 'validate-metrics.sh braucht periodischen Trigger'),
    ('External Signals', 'Stale (9+ Tage)', 'warn', 'Kein Auto-Refresh während Idle'),
    ('Learning Rate', '3.94/100 Tasks', 'ok', 'Stabil'),
]

for i, (metric, val, st, comment) in enumerate(kpis):
    row = r + 1 + i
    ws.cell(row=row, column=1, value=metric).font = BOLD
    status_cell(ws, row, 2, val, st)
    status_cell(ws, row, 3, 'OK' if st=='ok' else ('Achtung' if st=='warn' else 'Kritisch'), st)
    ws.cell(row=row, column=4, value=comment).font = NORMAL

# ============ Trend Window ============
r = r + 1 + len(kpis) + 1
style_section(ws, r, 4, 'Lernkurve (50-Task-Fenster)')
r += 1
style_header(ws, r, 4)
for c, h in enumerate(['Window', 'Success Rate', 'Timeouts', 'Trend'], 1):
    ws.cell(row=r, column=c, value=h)

windows = [
    ('1-50', '34%', 19, ''), ('51-100', '4%', 5, 'Tief'),
    ('101-200', '5%', 71, 'Tief'), ('201-300', '13%', 57, 'Langsamer Anstieg'),
    ('301-400', '13%', 17, 'Stabilisierung'), ('401-500', '24%', 49, 'Anstieg'),
    ('501-600', '12%', 31, 'Rückfall'), ('601-700', '72%', 2, 'Durchbruch'),
    ('701-786', '97%', 1, 'Peak'),
]
for i, (w, sr, to, trend) in enumerate(windows):
    row = r + 1 + i
    ws.cell(row=row, column=1, value=w).font = NORMAL
    st = 'ok' if float(sr.strip('%')) > 80 else ('warn' if float(sr.strip('%')) > 30 else 'bad')
    status_cell(ws, row, 2, sr, st)
    ws.cell(row=row, column=3, value=to).font = NORMAL
    ws.cell(row=row, column=4, value=trend).font = Font(name='Arial', size=10, italic=True)

# ============ SHEET 2: Tasks ============
ws2 = wb.create_sheet('Task Registry')
ws2.sheet_properties.tabColor = '2E75B6'
for c, w in enumerate([12, 55, 12, 15, 30], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

r = 1
style_header(ws2, r, 5)
for c, h in enumerate(['Status', 'Task', 'Attempts', 'Project', 'Bewertung'], 1):
    ws2.cell(row=r, column=c, value=h)

tasks = [
    ('completed', 'Add comment documenting MAX_STEP_CHARS=400', 3, 'local', 'Erledigt'),
    ('completed', 'Add test: planner output steps under 600 chars', 3, 'local', 'Erledigt'),
    ('completed', 'Update learner.sh dedup comment (65% threshold)', 3, 'local', 'Erledigt'),
    ('completed', 'Improve retry success rate', 1, 'local', 'Erledigt'),
    ('shelved', 'Unit test: clamp_prompt_context 4000-char limit', 2, 'local', 'Zu oft fehlgeschlagen'),
    ('shelved', 'Unit test: classify_retry_failure categories', 2, 'local', 'Zu oft fehlgeschlagen'),
    ('shelved', 'Update learner.sh dedup threshold comment', 2, 'local', 'Duplikat — bereits erledigt'),
    ('shelved', 'Review external signal: OpenAI Python v2.30.0', 0, 'local', 'Signal veraltet'),
    ('shelved', 'Keep executable system-work buffer', 0, 'local', 'Nicht mehr relevant'),
    ('shelved', 'Inventory: cap pre-step planning budget', 0, 'local', 'Aufgeschoben'),
    ('shelved', 'Reduce timeout rate (47%)', 0, 'local', 'Bereits auf ~2% gefallen'),
    ('completed', 'Inventory: verify dashboard incident id field', '-', 'superheld', 'Erledigt'),
    ('completed', 'Verify trigger-aware credential recovery routing', '-', 'superheld', 'Erledigt'),
    ('completed', 'Fix value measurement blindness', '-', 'superheld', 'Teilweise gelöst (score=5)'),
    ('shelved', 'Inventory: fix value measurement blindness', '-', 'superheld', 'Timeout bei Step 3'),
]

for i, (status, title, att, proj, bew) in enumerate(tasks):
    row = r + 1 + i
    st = 'ok' if status == 'completed' else 'warn'
    status_cell(ws2, row, 1, status.upper(), st)
    ws2.cell(row=row, column=2, value=title).font = NORMAL
    ws2.cell(row=row, column=3, value=att).font = NORMAL
    ws2.cell(row=row, column=4, value=proj).font = NORMAL
    ws2.cell(row=row, column=5, value=bew).font = NORMAL

# ============ SHEET 3: Handlungsempfehlungen ============
ws3 = wb.create_sheet('Empfehlungen')
ws3.sheet_properties.tabColor = 'ED7D31'
for c, w in enumerate([8, 50, 15, 50], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

r = 1
ws3.merge_cells('A1:D1')
ws3.cell(row=1, column=1, value='Handlungsempfehlungen & Modifikationen').font = Font(name='Arial', bold=True, size=13, color='1F4E79')

r = 3
style_header(ws3, r, 4)
for c, h in enumerate(['#', 'Maßnahme', 'Priorität', 'Begründung'], 1):
    ws3.cell(row=r, column=c, value=h)

recommendations = [
    (1, 'validate-metrics.sh periodisch triggern (alle 6h)', 'HOCH',
     'Metrics driften während Idle-Perioden — 8 Audits haben dasselbe Problem gefunden und manuell gefixt'),
    (2, 'Rule-Eviction-Mechanismus einbauen', 'HOCH',
     'Rules sind bei 20/20 am Cap. System kann keine neuen Regeln lernen. 5 Kandidaten warten.'),
    (3, 'Growth-Mode in self-improve.sh implementieren', 'MITTEL',
     'Pipeline idle + 98% Success → System sollte proaktiv Test-Coverage, Doku, Observability erweitern'),
    (4, 'External Signals Auto-Refresh', 'MITTEL',
     'Signals seit 9+ Tagen stale. Dependency-Updates werden verpasst.'),
    (5, 'Evaluator Score-Kalibrierung', 'MITTEL',
     '24% der Tasks bekommen score=0 — Evaluator unterscheidet nicht zwischen produktiv und Verschwendung'),
    (6, 'Shelved Tasks reviewen und aufräumen', 'NIEDRIG',
     '7 shelved Tasks — einige sind Duplikate oder bereits gelöst (z.B. "Reduce timeout rate" ist faktisch erledigt)'),
]

for i, (num, action, prio, reason) in enumerate(recommendations):
    row = r + 1 + i
    ws3.cell(row=row, column=1, value=num).font = BOLD
    ws3.cell(row=row, column=2, value=action).font = BOLD
    p_st = 'bad' if prio == 'HOCH' else ('warn' if prio == 'MITTEL' else 'ok')
    status_cell(ws3, row, 3, prio, p_st)
    ws3.cell(row=row, column=4, value=reason).font = NORMAL
    ws3.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

# Verdict section
r = r + 1 + len(recommendations) + 2
style_section(ws3, r, 4, 'Gesamtbewertung')
r += 1
ws3.merge_cells(f'A{r}:D{r}')
ws3.cell(row=r, column=1, value='Tasks umsetzbar: JA — die bestehenden Tasks sind korrekt kategorisiert. '
    '4 completed, 7 shelved (davon 3 korrekt pausiert, 2 Duplikate, 2 nicht mehr relevant).').font = NORMAL
ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws3.row_dimensions[r].height = 35

r += 1
ws3.merge_cells(f'A{r}:D{r}')
ws3.cell(row=r, column=1, value='Success Rate: STABIL bei 98% — keine Regression. '
    'All-time 31% wird durch frühe Lernphase gedrückt, nicht durch aktuelle Probleme.').font = NORMAL
ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws3.row_dimensions[r].height = 35

r += 1
ws3.merge_cells(f'A{r}:D{r}')
ws3.cell(row=r, column=1, value='Modifikationen: 2 HOCH-Prioritäts-Änderungen empfohlen (Metrics-Validation-Scheduling, Rule-Eviction). '
    'Ohne diese stagniert das System trotz guter Performance — es kann nicht weiterlernen und Metriken driften.').font = Font(name='Arial', size=10, bold=True, color='9C0006')
ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws3.row_dimensions[r].height = 50

outpath = '/sessions/pensive-laughing-heisenberg/mnt/codex-agent-system/fortschrittsbericht-2026-04-03-status-check.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')
