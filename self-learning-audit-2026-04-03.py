from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# --- SUMMARY SHEET ---
ws = wb.active
ws.title = "Audit Summary"

header_font = Font(bold=True, size=14, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
section_font = Font(bold=True, size=12, color="2F5496")
good_fill = PatternFill("solid", fgColor="C6EFCE")
bad_fill = PatternFill("solid", fgColor="FFC7CE")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, cols, font=header_font, fill=header_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

ws.merge_cells("A1:F1")
ws["A1"] = "Self-Learning Audit Report — 2026-04-03"
ws["A1"].font = Font(bold=True, size=16, color="2F5496")
ws["A1"].alignment = Alignment(horizontal="center")

ws["A3"] = "Frage: Lernt das System effizient dazu? Wird es bei jeder Iteration messbar besser?"
ws["A3"].font = Font(italic=True, size=11)

# Overall verdict
ws["A5"] = "Gesamtbewertung"
ws["A5"].font = section_font
ws["B5"] = "JA — System lernt messbar dazu, aber mit Engpässen"
ws["B5"].font = Font(bold=True, size=11)
ws["B5"].fill = warn_fill

# Key metrics table
r = 7
headers = ["Metrik", "Wert", "Trend", "Bewertung"]
for i, h in enumerate(headers, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)

metrics = [
    ("All-time Success Rate", "31%", "Baseline", ""),
    ("Recent Success Rate (last 50)", "98%", "+67pp vs all-time", "good"),
    ("First-Pass Success Rate", "80%", "Stabil", "good"),
    ("Timeout Rate (all-time)", "27%", "Sinkend (0 in letzten 50)", "good"),
    ("Learning Rules Count", "21 (war 14)", "+50% gewachsen", "good"),
    ("Learning Rate", "2.8/100 Tasks", "Verbessert (war 0.51)", "good"),
    ("Retry Classification Coverage", "100%", "War 24%", "good"),
    ("Self-Improve Task Success", "0%", "Blockiert", "bad"),
    ("External Signals", "Stale (8 Tage)", "Refresh nötig", "bad"),
    ("Zombie Tasks", "20 (166 Slots verschwendet)", "Unverändert", "warn"),
    ("Loop Effort", "11 Tasks, 26 Extra Steps", "Historisch (gelöst)", "warn"),
    ("code_quality Success", "<11% (beide Provider)", "Keine Verbesserung", "bad"),
    ("learning Category Success", "<12% (beide Provider)", "Keine Verbesserung", "bad"),
]

for i, (m, v, t, rating) in enumerate(metrics):
    row = r + 1 + i
    ws.cell(row=row, column=1, value=m)
    ws.cell(row=row, column=2, value=v)
    ws.cell(row=row, column=3, value=t)
    fill_map = {"good": good_fill, "bad": bad_fill, "warn": warn_fill}
    label_map = {"good": "OK", "bad": "Problem", "warn": "Warnung", "": "—"}
    cell = ws.cell(row=row, column=4, value=label_map.get(rating, "—"))
    if rating in fill_map:
        cell.fill = fill_map[rating]
    style_row(ws, row, 4)

for c in [20, 30, 25, 12]:
    ws.column_dimensions[get_column_letter(metrics.index(metrics[0])+1)].width = 20
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 15

# --- TREND SHEET ---
ws2 = wb.create_sheet("Iteration Trend")
ws2["A1"] = "Success Rate by 50-Task Windows"
ws2["A1"].font = section_font

windows = [
    ("1-50", 0.34, 19), ("51-100", 0.04, 5), ("101-150", 0.06, 33),
    ("151-200", 0.04, 38), ("201-250", 0.16, 34), ("251-300", 0.10, 23),
    ("301-350", 0.14, 5), ("351-400", 0.12, 12), ("401-450", 0.22, 29),
    ("451-500", 0.26, 20), ("501-550", 0.10, 23), ("551-600", 0.14, 8),
    ("601-650", 0.58, 1), ("651-700", 0.86, 1), ("701-750", 0.96, 0),
    ("751-785", 0.97, 1),
]

for i, h in enumerate(["Window", "Success Rate", "Timeouts", "Visual"], 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 4)

for i, (w, rate, to) in enumerate(windows):
    row = 4 + i
    ws2.cell(row=row, column=1, value=w)
    ws2.cell(row=row, column=2, value=rate)
    ws2.cell(row=row, column=2).number_format = "0%"
    ws2.cell(row=row, column=3, value=to)
    bar = "█" * int(rate * 20)
    ws2.cell(row=row, column=4, value=bar)
    if rate >= 0.8:
        ws2.cell(row=row, column=4).font = Font(color="006100")
    elif rate >= 0.3:
        ws2.cell(row=row, column=4).font = Font(color="9C6500")
    else:
        ws2.cell(row=row, column=4).font = Font(color="9C0006")
    style_row(ws2, row, 4)

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 25

# --- PROBLEMS & FIXES SHEET ---
ws3 = wb.create_sheet("Probleme & Fixes")
ws3["A1"] = "Identifizierte Probleme und durchgeführte Fixes"
ws3["A1"].font = section_font

for i, h in enumerate(["#", "Problem", "Schwere", "Fix durchgeführt", "Datei"], 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 5)

fixes = [
    (1, "Self-improve Tasks haben 0% Erfolg — System kann sich nicht selbst verbessern", "Kritisch",
     "Neue Regel: Self-improve auf Config/Data-Files beschränken, keine Shell-Script-Änderungen", "rules.md, prompt-rules.md, CLAUDE.md"),
    (2, "code_quality & learning Kategorien <12% Erfolg bei beiden Providern", "Hoch",
     "Neue Routing-Regel: split_required flag, Aufteilen in read-only Analyse + Write-Task", "provider-routing.json, rules.md"),
    (3, "External Signals stale seit 8 Tagen", "Mittel",
     "Neue Regel: Stale-Signal-Refresh bei jeder Metrics-Berechnung triggern", "rules.md"),
    (4, "Loop Effort & Retry Churn Alerts aktiv obwohl recent_success_rate 0.98", "Mittel",
     "Neue Regel: Stale Alerts auto-clearen wenn recent_success_rate >= 0.95", "rules.md, CLAUDE.md"),
    (5, "title_family_cooldown blockiert alle Self-Improve Verbesserungen", "Hoch",
     "Neue Regel: Cooldown auf max 24h begrenzen für tägliche Improvement-Versuche", "rules.md, prompt-rules.md"),
    (6, "Dashboard-Verification Tasks scheitern an Review (runtime-path vs data-layer)", "Mittel",
     "Neue Regel: Dashboard-Tasks auf data-layer Assertions beschränken", "rules.md, prompt-rules.md"),
    (7, "CLAUDE.md enthält veraltete Metriken (Learning Rate 0.51 statt 2.8)", "Niedrig",
     "CLAUDE.md komplett aktualisiert mit korrekten Werten und neuen Rules", "CLAUDE.md"),
]

for i, (n, prob, sev, fix, files) in enumerate(fixes):
    row = 4 + i
    ws3.cell(row=row, column=1, value=n)
    ws3.cell(row=row, column=2, value=prob)
    sev_cell = ws3.cell(row=row, column=3, value=sev)
    if sev == "Kritisch": sev_cell.fill = bad_fill
    elif sev == "Hoch": sev_cell.fill = warn_fill
    ws3.cell(row=row, column=4, value=fix)
    ws3.cell(row=row, column=5, value=files)
    style_row(ws3, row, 5)

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 65
ws3.column_dimensions["E"].width = 30

# --- PROVIDER STATS SHEET ---
ws4 = wb.create_sheet("Provider Performance")
ws4["A1"] = "Provider Success Rates by Category"
ws4["A1"].font = section_font

for i, h in enumerate(["Category", "Codex Rate", "Codex Tasks", "Claude Rate", "Claude Tasks", "Winner", "Gap"], 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 7)

provdata = [
    ("auth", 0.724, 29, 0.0, 3, "codex", "+72.4pp"),
    ("ui", 0.406, 244, 0.182, 77, "codex", "+22.4pp"),
    ("infra", 0.402, 92, 0.156, 32, "codex", "+24.6pp"),
    ("testing", 0.571, 7, 0.357, 14, "codex", "+21.4pp"),
    ("general", 0.254, 142, 0.185, 54, "codex", "+6.9pp"),
    ("project", 0.133, 15, 0.0, 6, "codex", "+13.3pp"),
    ("learning", 0.118, 34, 0.083, 12, "codex", "+3.5pp"),
    ("code_quality", 0.105, 19, 0.0, 5, "codex", "+10.5pp"),
]

for i, (cat, cr, ct, clr, clt, winner, gap) in enumerate(provdata):
    row = 4 + i
    ws4.cell(row=row, column=1, value=cat)
    ws4.cell(row=row, column=2, value=cr); ws4.cell(row=row, column=2).number_format = "0.0%"
    ws4.cell(row=row, column=3, value=ct)
    ws4.cell(row=row, column=4, value=clr); ws4.cell(row=row, column=4).number_format = "0.0%"
    ws4.cell(row=row, column=5, value=clt)
    ws4.cell(row=row, column=6, value=winner)
    ws4.cell(row=row, column=7, value=gap)
    style_row(ws4, row, 7)

for c, w in zip("ABCDEFG", [15, 12, 12, 12, 12, 10, 10]):
    ws4.column_dimensions[c].width = w

outpath = "/sessions/festive-admiring-dijkstra/mnt/codex-agent-system/self-learning-audit-2026-04-03.xlsx"
wb.save(outpath)
print(f"Saved to {outpath}")
