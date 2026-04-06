from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color constants
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
RED_FONT = Font(name='Arial', bold=True, size=10, color='CC0000')
GREEN_FONT = Font(name='Arial', bold=True, size=10, color='006600')
ORANGE_FONT = Font(name='Arial', bold=True, size=10, color='CC6600')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_cell(ws, row, col, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align or Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border
    return cell

# ========== Sheet 1: Executive Summary ==========
ws = wb.active
ws.title = 'Executive Summary'
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50

ws.cell(row=1, column=1, value='Codex Agent System — Fortschrittsbericht').font = TITLE_FONT
ws.cell(row=2, column=1, value='Datum: 2026-04-04 | Pipeline: IDLE seit ~69h').font = Font(name='Arial', size=10, italic=True, color='666666')

row = 4
style_header_row(ws, row, 3)
ws.cell(row=row, column=1, value='Metrik')
ws.cell(row=row, column=2, value='Wert')
ws.cell(row=row, column=3, value='Bewertung')

metrics = [
    ('Total Tasks (traced)', '312 / 786', 'Solide Datenbasis'),
    ('All-time Success Rate (trace)', '59.6%', 'Durch frühe Fehler gedrückt'),
    ('Recent-50 Success Rate', '0%', 'KRITISCH — letzten 50 alle FAIL (score=0 meta-Tasks)'),
    ('Effective Success Rate (score>0)', '6%', 'KRITISCH — fast kein produktiver Output'),
    ('Meta-Task Ratio (all)', '79.2%', 'Zu hoch — System introspektiert statt zu produzieren'),
    ('Meta-Task Ratio (recent 50)', '100%', 'KRITISCH — nur noch Nabelschau'),
    ('Pipeline Status', 'IDLE seit 69h', 'Braucht Host-Run zur Validierung v33-v36 Fixes'),
    ('Queue', '0 Tasks', 'Leer — kein Nachschub'),
    ('Registry Größe', '125 KB / 512 KB', 'Gesund, kein Druck'),
    ('Aktive Regeln', '39 (20 + 19)', 'Am Limit, Eviction funktioniert'),
    ('Zombie Tasks', '20', '166 verschwendete Slots'),
    ('Letzter produktiver Task', '2026-03-31', '4 Tage ohne echten Output'),
]

for i, (m, v, b) in enumerate(metrics):
    r = row + 1 + i
    style_cell(ws, r, 1, BOLD_FONT)
    ws.cell(row=r, column=1, value=m)
    c2 = style_cell(ws, r, 2, BOLD_FONT)
    ws.cell(row=r, column=2, value=v)
    style_cell(ws, r, 3, BODY_FONT)
    ws.cell(row=r, column=3, value=b)
    if 'KRITISCH' in b:
        c2.font = RED_FONT
        ws.cell(row=r, column=3).font = RED_FONT
        ws.cell(row=r, column=3).fill = RED_FILL
    elif 'Gesund' in b or 'funktioniert' in b:
        ws.cell(row=r, column=3).fill = GREEN_FILL
    if i % 2 == 0:
        for cc in range(1, 4):
            if not ws.cell(row=r, column=cc).fill or ws.cell(row=r, column=cc).fill.fgColor.rgb == '00000000':
                ws.cell(row=r, column=cc).fill = LIGHT_GRAY

# ========== Sheet 2: Trend-Verlauf ==========
ws2 = wb.create_sheet('Trend-Verlauf')
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 50

ws2.cell(row=1, column=1, value='Success Rate Trend (50er-Fenster)').font = TITLE_FONT
row = 3
style_header_row(ws2, row, 4)
ws2.cell(row=row, column=1, value='Task-Fenster')
ws2.cell(row=row, column=2, value='Success Rate')
ws2.cell(row=row, column=3, value='Timeouts')
ws2.cell(row=row, column=4, value='Visuell')

windows = [
    ('1-50', 0.34, 19), ('51-100', 0.04, 5), ('101-150', 0.06, 33),
    ('151-200', 0.04, 38), ('201-250', 0.16, 34), ('251-300', 0.10, 23),
    ('301-350', 0.14, 5), ('351-400', 0.12, 12), ('401-450', 0.22, 29),
    ('451-500', 0.26, 20), ('501-550', 0.10, 23), ('551-600', 0.14, 8),
    ('601-650', 0.58, 1), ('651-700', 0.86, 1), ('701-750', 0.96, 0),
    ('751-786', 0.97, 1),
]

for i, (w, rate, to) in enumerate(windows):
    r = row + 1 + i
    style_cell(ws2, r, 1, BODY_FONT)
    ws2.cell(row=r, column=1, value=w)
    c = style_cell(ws2, r, 2, BOLD_FONT)
    ws2.cell(row=r, column=2, value=rate)
    c.number_format = '0%'
    if rate >= 0.90:
        c.fill = GREEN_FILL; c.font = GREEN_FONT
    elif rate >= 0.50:
        c.fill = YELLOW_FILL; c.font = ORANGE_FONT
    else:
        c.fill = RED_FILL; c.font = RED_FONT
    style_cell(ws2, r, 3, BODY_FONT)
    ws2.cell(row=r, column=3, value=to)
    style_cell(ws2, r, 4, BODY_FONT)
    bar = '█' * int(rate * 20) + '░' * (20 - int(rate * 20))
    ws2.cell(row=r, column=4, value=bar)

r = row + len(windows) + 2
ws2.cell(row=r, column=1, value='ACHTUNG').font = RED_FONT
ws2.cell(row=r, column=2, value='Die 97% Success Rate der letzten Fenster ist irreführend!').font = RED_FONT
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
ws2.cell(row=r, column=1, value='').font = BODY_FONT
ws2.cell(row=r, column=2, value='Die letzten 50 Trace-Einträge zeigen 0 Successes, 50 Failures (alles Meta-Tasks mit score=0).').font = BODY_FONT
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
ws2.cell(row=r, column=2, value='Effektive Wertproduktion seit 31.03.: NULL. System produziert nur Inventory/Verify-Loops.').font = ORANGE_FONT
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

# ========== Sheet 3: Diagnose & Empfehlungen ==========
ws3 = wb.create_sheet('Diagnose & Empfehlungen')
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 14

ws3.cell(row=1, column=1, value='Diagnose der aktuellen Probleme').font = TITLE_FONT

row = 3
style_header_row(ws3, row, 4)
ws3.cell(row=row, column=1, value='#')
ws3.cell(row=row, column=2, value='Problem')
ws3.cell(row=row, column=3, value='Details')
ws3.cell(row=row, column=4, value='Schwere')

problems = [
    ('Meta-Task Spirale', 'Letzte 50 Tasks sind 100% Meta (Inventory/Verify). Kein produktiver Task seit 31.03. System dreht sich im Kreis.', 'KRITISCH'),
    ('Score=0 Epidemie', '94% der "Successes" haben Score=0. Effective Success Rate nur 6%. Die 98% Headline-Rate maskiert null Wertproduktion.', 'KRITISCH'),
    ('Growth-Mode Deadlock (v36 Fix)', 'Meta-Ratio Guard blockierte ALLE Task-Generation inkl. Growth-Mode. Triple-Deadlock in self-improve.sh. v36 Fix vorhanden aber UNGETESTET.', 'HOCH'),
    ('Pipeline seit 69h IDLE', 'Kein Host-Prozess führt self-improve.sh aus. 8 Code-Fixes (v33-v36) warten auf Validierung. idle-heartbeat.sh existiert aber nicht scheduled.', 'HOCH'),
    ('Zombie Tasks', '20 Zombie-Tasks haben 166 Execution-Slots verschwendet. Gleiche Tasks werden wiederholt ausgeführt trotz permanentem Scheitern.', 'MITTEL'),
]

for i, (p, d, s) in enumerate(problems):
    r = row + 1 + i
    style_cell(ws3, r, 1, BOLD_FONT); ws3.cell(row=r, column=1, value=i+1)
    style_cell(ws3, r, 2, BOLD_FONT); ws3.cell(row=r, column=2, value=p)
    style_cell(ws3, r, 3, BODY_FONT); ws3.cell(row=r, column=3, value=d)
    sc = style_cell(ws3, r, 4, BOLD_FONT); ws3.cell(row=r, column=4, value=s)
    sc.alignment = Alignment(horizontal='center')
    if s == 'KRITISCH':
        sc.fill = RED_FILL; sc.font = RED_FONT
    elif s == 'HOCH':
        sc.fill = YELLOW_FILL; sc.font = ORANGE_FONT

row = row + len(problems) + 3
ws3.cell(row=row, column=1, value='Empfohlene Maßnahmen').font = TITLE_FONT

row += 2
style_header_row(ws3, row, 4)
ws3.cell(row=row, column=1, value='Prio')
ws3.cell(row=row, column=2, value='Maßnahme')
ws3.cell(row=row, column=3, value='Details')
ws3.cell(row=row, column=4, value='Aufwand')

actions = [
    ('1', 'Host-Run: self-improve.sh ausführen', 'v33-v36 Fixes validieren. Ohne Host-Run kann kein einziger Fix wirken. Meta-Ratio Guard + Growth-Mode Deadlock Fix müssen live getestet werden.', 'Gering'),
    ('2', 'idle-heartbeat.sh per launchd/cron schedulen', 'Alle 6h memory-sync.sh aufrufen. Verhindert 69h+ Idle-Phasen. Script existiert bereits.', 'Gering'),
    ('3', 'Manuell 2-3 produktive Tasks einspeisen', 'Meta-Ratio von 100% kann nur sinken wenn produktive Tasks laufen. Vorschlag: echte Code-Tasks für superheld-Projekt.', 'Mittel'),
    ('4', 'Evaluator-Scoring validieren nach v32 Fix', 'v32 hat Scoring-Rubrik geändert (Introspection=1-2, Code=5-10). Muss mit realen Tasks validiert werden.', 'Mittel'),
    ('5', 'Trace-Daten bereinigen', 'Die 50 FAIL/score=0 Einträge am Ende des Traces verzerren alle Metriken. Ggf. archivieren und Baseline neu setzen.', 'Gering'),
]

for i, (p, m, d, a) in enumerate(actions):
    r = row + 1 + i
    pc = style_cell(ws3, r, 1, BOLD_FONT); ws3.cell(row=r, column=1, value=p)
    pc.alignment = Alignment(horizontal='center')
    style_cell(ws3, r, 2, BOLD_FONT); ws3.cell(row=r, column=2, value=m)
    style_cell(ws3, r, 3, BODY_FONT); ws3.cell(row=r, column=3, value=d)
    ac = style_cell(ws3, r, 4, BODY_FONT); ws3.cell(row=r, column=4, value=a)
    ac.alignment = Alignment(horizontal='center')

# ========== Sheet 4: Fazit ==========
ws4 = wb.create_sheet('Fazit')
ws4.column_dimensions['A'].width = 80

ws4.cell(row=1, column=1, value='Gesamtbewertung').font = TITLE_FONT

ws4.cell(row=3, column=1, value='Sind die bisherigen Tasks umsetzbar?').font = BOLD_FONT
ws4.cell(row=4, column=1, value='JEIN. Die Tasks selbst sind technisch umsetzbar (Trend 34%→97% zeigt funktionierendes Lernsystem). ABER: Das System produziert seit 4 Tagen keinen echten Output. 100% der letzten 50 Tasks sind Meta/Introspection-Tasks. Die hohe Success Rate ist eine Illusion — effektiver Wert nahe null.').font = BODY_FONT
ws4.cell(row=4, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[4].height = 55

ws4.cell(row=6, column=1, value='Heben wir die Success Rate?').font = BOLD_FONT
ws4.cell(row=7, column=1, value='Die Headline-Rate (98%) ist irreführend. Die echte Effective Rate (score>0) liegt bei 6%. Verbesserung erfordert: (1) v36 Growth-Mode Fix validieren per Host-Run, (2) produktive Tasks einspeisen um Meta-Ratio unter 40% zu drücken, (3) Evaluator-Scoring nach v32 Fix prüfen.').font = BODY_FONT
ws4.cell(row=7, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[7].height = 55

ws4.cell(row=9, column=1, value='Sind Modifikationen notwendig?').font = BOLD_FONT
ws4.cell(row=10, column=1, value='JA — aber nicht an den Tasks oder Regeln, sondern an der Infrastruktur:').font = RED_FONT
ws4.cell(row=11, column=1, value='1. DRINGEND: Host-Run von self-improve.sh um v33-v36 Code-Fixes zu aktivieren (Meta-Ratio Guard, Growth-Mode Deadlock Fix, Evaluator Fix, Family Ceiling Fix, Freeze Guard, Cooldown Fix)').font = BODY_FONT
ws4.cell(row=11, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[11].height = 40
ws4.cell(row=12, column=1, value='2. DRINGEND: idle-heartbeat.sh schedulen (launchd/cron) — ohne dies bleibt Pipeline nach jedem Idle permanent stehen').font = BODY_FONT
ws4.cell(row=12, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[12].height = 30
ws4.cell(row=13, column=1, value='3. EMPFOHLEN: 2-3 manuelle produktive Tasks einspeisen um Meta-Ratio-Deadlock von der Daten-Seite aufzubrechen').font = BODY_FONT
ws4.cell(row=13, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[13].height = 30

ws4.cell(row=15, column=1, value='Positiv: Das Regelsystem (39 Regeln), die Knowledge Base (219 Einträge), und der Retry-Klassifikator (100% Coverage) sind in gutem Zustand. Die Code-Fixes v33-v36 adressieren alle bekannten Probleme — sie müssen nur aktiviert werden.').font = GREEN_FONT
ws4.cell(row=15, column=1).alignment = Alignment(wrap_text=True)
ws4.row_dimensions[15].height = 55

outpath = '/sessions/keen-sweet-lovelace/mnt/codex-agent-system/fortschrittsbericht-2026-04-04-scheduled-system-v3.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')
